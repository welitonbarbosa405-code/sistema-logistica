from flask import Blueprint, render_template, request, jsonify, flash, redirect, url_for, make_response
from flask_login import login_required, current_user
from app import db
from app.models import Consumivel, MovimentacaoEstoque
from app.utils.decorators import admin_required
from datetime import datetime
import logging
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

def normalizar_unidade(unidade_valor: str) -> str:
    """Normaliza e garante um valor padrão para unidade de medida."""
    if unidade_valor is None:
        return 'UN'
    unidade = unidade_valor.strip().upper()
    return unidade or 'UN'

# Criar o blueprint aqui
consumiveis_bp = Blueprint('consumiveis', __name__, url_prefix='/consumiveis')

@consumiveis_bp.route('/')
@login_required
def index():
    """Página principal de consumíveis"""
    try:
        # Buscar consumíveis com paginação
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 200, type=int)  # Aumentado para mostrar mais registros
        
        # Filtros
        filial_filtro = request.args.get('filial', '')
        nivel_filtro = request.args.get('nivel', '')
        
        query = Consumivel.query
        
        if filial_filtro:
            query = query.filter(Consumivel.filial.ilike(f'%{filial_filtro}%'))
        
        if nivel_filtro:
            query = query.filter(Consumivel.nivel_estoque == nivel_filtro)
        
        consumiveis = query.order_by(Consumivel.descricao.asc()).paginate(
            page=page, per_page=per_page, error_out=False
        )
        
        # Estatísticas
        total_consumiveis = Consumivel.query.count()
        estoque_critico = Consumivel.query.filter(Consumivel.nivel_estoque == 'Crítico').count()
        estoque_baixo = Consumivel.query.filter(Consumivel.nivel_estoque == 'Baixo').count()
        estoque_esgotado = Consumivel.query.filter(Consumivel.nivel_estoque == 'Esgotado').count()
        
        # Filiais únicas para filtro
        filiais = db.session.query(Consumivel.filial).distinct().all()
        filiais = [f[0] for f in filiais]
        
        return render_template('consumiveis_index.html',
                             consumiveis=consumiveis,
                             total_consumiveis=total_consumiveis,
                             estoque_critico=estoque_critico,
                             estoque_baixo=estoque_baixo,
                             estoque_esgotado=estoque_esgotado,
                             filiais=filiais,
                             filial_filtro=filial_filtro,
                             nivel_filtro=nivel_filtro)
    
    except Exception as e:
        logger.error(f"Erro ao carregar página de consumíveis: {str(e)}")
        flash('Erro ao carregar dados dos consumíveis', 'error')
        return render_template('consumiveis_index.html', consumiveis=None)

@consumiveis_bp.route('/api/consumiveis')
@login_required
def api_consumiveis():
    """API para buscar consumíveis"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 200, type=int)  # Aumentado para mostrar mais registros
        
        # Filtros
        filial = request.args.get('filial', '')
        nivel = request.args.get('nivel', '')
        codigo = request.args.get('codigo', '')
        
        query = Consumivel.query
        
        if filial:
            query = query.filter(Consumivel.filial.ilike(f'%{filial}%'))
        if nivel:
            query = query.filter(Consumivel.nivel_estoque == nivel)
        if codigo:
            query = query.filter(Consumivel.codigo.ilike(f'%{codigo}%'))
        
        consumiveis = query.order_by(Consumivel.descricao.asc()).paginate(
            page=page, per_page=per_page, error_out=False
        )
        
        return jsonify({
            'consumiveis': [c.to_dict() for c in consumiveis.items],
            'total': consumiveis.total,
            'pages': consumiveis.pages,
            'current_page': consumiveis.page,
            'has_next': consumiveis.has_next,
            'has_prev': consumiveis.has_prev
        })
    
    except Exception as e:
        logger.error(f"Erro na API de consumíveis: {str(e)}")
        return jsonify({'error': f'Erro ao buscar consumíveis: {str(e)}'}), 500

@consumiveis_bp.route('/teste', methods=['GET'])
def teste():
    """Rota de teste para verificar se o modelo está funcionando"""
    try:
        # Testar se consegue criar um objeto Consumivel
        consumivel_teste = Consumivel(
            filial='TESTE',
            codigo='TESTE001',
            descricao='Consumível de teste',
            saldo_estoque=10,
            estoque_minimo=5,
            unidade_medida='UN'
        )
        
        # Testar se consegue calcular nível de estoque
        nivel = consumivel_teste.calcular_nivel_estoque()
        
        return jsonify({
            'success': True,
            'message': 'Modelo Consumivel funcionando corretamente',
            'nivel_calculado': nivel,
            'consumivel_teste': consumivel_teste.to_dict()
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Erro no teste: {str(e)}',
            'tipo_erro': type(e).__name__
        }), 500

@consumiveis_bp.route('/teste-cadastro', methods=['POST'])
def teste_cadastro():
    """Rota de teste para cadastro sem decorators"""
    try:
        logger.info("Iniciando teste de cadastro")
        data = request.get_json() if request.is_json else request.form
        logger.info(f"Dados recebidos: {data}")
        
        filial = data.get('filial', 'TESTE').strip()
        codigo = data.get('codigo', 'TESTE002').strip()
        descricao = data.get('descricao', 'Teste de cadastro').strip()
        saldo_estoque = int(data.get('saldo_estoque', 0))
        estoque_minimo = int(data.get('estoque_minimo', 0))
        unidade_medida = normalizar_unidade(data.get('unidade_medida', 'UN'))
        
        logger.info(f"Valores processados - Filial: {filial}, Código: {codigo}")
        
        # Criar novo consumível
        logger.info("Criando novo objeto Consumivel")
        consumivel = Consumivel(
            filial=filial,
            codigo=codigo,
            descricao=descricao,
            saldo_estoque=saldo_estoque,
            estoque_minimo=estoque_minimo,
            unidade_medida=unidade_medida
        )
        
        # Calcular nível de estoque
        logger.info("Calculando nível de estoque")
        consumivel.atualizar_nivel_estoque()
        
        logger.info("Adicionando ao banco de dados")
        db.session.add(consumivel)
        db.session.commit()
        
        logger.info(f"Consumível {codigo} cadastrado com sucesso!")
        
        return jsonify({
            'success': True,
            'message': 'Teste de cadastro realizado com sucesso!',
            'consumivel': consumivel.to_dict()
        })
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro no teste de cadastro: {str(e)}")
        logger.error(f"Tipo do erro: {type(e).__name__}")
        
        # Verificar se é erro de constraint única
        if 'UNIQUE constraint failed' in str(e) and 'filial' in str(e) and 'codigo' in str(e):
            return jsonify({
                'error': f'⚠️ Consumível já cadastrado!',
                'details': f'O código "{codigo}" já existe para a filial "{filial}".',
                'suggestion': 'Verifique se o código está correto ou edite o consumível existente.'
            }), 400
        
        return jsonify({'error': f'Erro no teste: {str(e)}'}), 500

@consumiveis_bp.route('/cadastrar', methods=['GET', 'POST'])
@login_required
def cadastrar():
    """Cadastrar novo consumível"""
    if request.method == 'POST':
        try:
            logger.info("Iniciando cadastro de consumível")
            data = request.get_json() if request.is_json else request.form
            logger.info(f"Dados recebidos: {data}")
            
            filial = data.get('filial', '').strip()
            codigo = data.get('codigo', '').strip()
            descricao = data.get('descricao', '').strip()
            saldo_estoque = int(data.get('saldo_estoque', 0))
            estoque_minimo = int(data.get('estoque_minimo', 0))
            unidade_medida = normalizar_unidade(data.get('unidade_medida', 'UN'))
            
            logger.info(f"Valores processados - Filial: {filial}, Código: {codigo}, Descrição: {descricao}")
            
            # Validações
            if not all([filial, codigo, descricao, unidade_medida]):
                logger.warning("Campos obrigatórios não preenchidos")
                return jsonify({'error': 'Filial, código, descrição e unidade de medida são obrigatórios'}), 400
            
            if len(unidade_medida) > 20:
                logger.warning("Unidade de medida excedeu limite máximo")
                return jsonify({'error': 'Unidade de medida deve ter no máximo 20 caracteres'}), 400
            
            if saldo_estoque < 0 or estoque_minimo < 0:
                logger.warning("Valores de estoque negativos")
                return jsonify({'error': 'Valores de estoque não podem ser negativos'}), 400
            
            # Verificar se já existe (constraint única)
            logger.info("Verificando se consumível já existe")
            existente = Consumivel.query.filter_by(filial=filial, codigo=codigo).first()
            if existente:
                logger.warning(f"Consumível {codigo} já existe para filial {filial}")
                return jsonify({
                    'error': f'⚠️ Consumível já cadastrado!',
                    'details': f'O código "{codigo}" já existe para a filial "{filial}".',
                    'suggestion': 'Verifique se o código está correto ou edite o consumível existente.',
                    'existing_item': {
                        'id': existente.id,
                        'descricao': existente.descricao,
                        'saldo_estoque': existente.saldo_estoque,
                        'nivel_estoque': existente.nivel_estoque
                    }
                }), 400
            
            # Criar novo consumível
            logger.info("Criando novo objeto Consumivel")
            consumivel = Consumivel(
                filial=filial,
                codigo=codigo,
                descricao=descricao,
                saldo_estoque=saldo_estoque,
                estoque_minimo=estoque_minimo,
                unidade_medida=unidade_medida
            )
            
            # Calcular nível de estoque
            logger.info("Calculando nível de estoque")
            consumivel.atualizar_nivel_estoque()
            
            logger.info("Adicionando ao banco de dados")
            db.session.add(consumivel)
            db.session.commit()
            
            logger.info(f"Consumível {codigo} cadastrado com sucesso para filial {filial} por {current_user.email}")
            
            return jsonify({
                'success': True,
                'message': 'Consumível cadastrado com sucesso!',
                'consumivel': consumivel.to_dict()
            })
        
        except ValueError as e:
            logger.error(f"Erro de valor: {str(e)}")
            return jsonify({'error': 'Valores numéricos inválidos'}), 400
        except Exception as e:
            db.session.rollback()
            logger.error(f"Erro ao cadastrar consumível: {str(e)}")
            logger.error(f"Tipo do erro: {type(e).__name__}")
            
            # Verificar se é erro de constraint única
            if 'UNIQUE constraint failed' in str(e) and 'filial' in str(e) and 'codigo' in str(e):
                return jsonify({
                    'error': f'⚠️ Consumível já cadastrado!',
                    'details': f'O código "{codigo}" já existe para a filial "{filial}".',
                    'suggestion': 'Verifique se o código está correto ou edite o consumível existente.'
                }), 400
            
            return jsonify({'error': f'Erro ao cadastrar consumível: {str(e)}'}), 500
    
    # GET - mostrar formulário
    return render_template('consumiveis_cadastrar.html')

@consumiveis_bp.route('/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar(id):
    """Editar consumível existente"""
    consumivel = Consumivel.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            data = request.get_json() if request.is_json else request.form
            
            # Atualizar campos
            consumivel.filial = data.get('filial', consumivel.filial).strip()
            consumivel.codigo = data.get('codigo', consumivel.codigo).strip()
            consumivel.descricao = data.get('descricao', consumivel.descricao).strip()
            consumivel.saldo_estoque = int(data.get('saldo_estoque', consumivel.saldo_estoque))
            consumivel.estoque_minimo = int(data.get('estoque_minimo', consumivel.estoque_minimo))
            consumivel.unidade_medida = normalizar_unidade(
                data.get('unidade_medida', consumivel.unidade_medida)
            )
            
            # Validações
            if not all([consumivel.filial, consumivel.codigo, consumivel.descricao, consumivel.unidade_medida]):
                return jsonify({'error': 'Filial, código, descrição e unidade de medida são obrigatórios'}), 400
            
            if len(consumivel.unidade_medida) > 20:
                return jsonify({'error': 'Unidade de medida deve ter no máximo 20 caracteres'}), 400
            
            if consumivel.saldo_estoque < 0 or consumivel.estoque_minimo < 0:
                return jsonify({'error': 'Valores de estoque não podem ser negativos'}), 400
            
            # Verificar constraint única (exceto o próprio registro)
            existente = Consumivel.query.filter(
                Consumivel.filial == consumivel.filial,
                Consumivel.codigo == consumivel.codigo,
                Consumivel.id != id
            ).first()
            
            if existente:
                return jsonify({'error': f'Consumível {consumivel.codigo} já cadastrado para a filial {consumivel.filial}'}), 400
            
            # Atualizar nível de estoque
            consumivel.atualizar_nivel_estoque()
            
            db.session.commit()
            
            logger.info(f"Consumível {consumivel.codigo} editado por {current_user.email}")
            
            return jsonify({
                'success': True,
                'message': 'Consumível atualizado com sucesso!',
                'consumivel': consumivel.to_dict()
            })
        
        except ValueError as e:
            return jsonify({'error': 'Valores numéricos inválidos'}), 400
        except Exception as e:
            db.session.rollback()
            logger.error(f"Erro ao editar consumível: {str(e)}")
            return jsonify({'error': f'Erro ao editar consumível: {str(e)}'}), 500
    
    # GET - mostrar formulário de edição
    return render_template('consumiveis_editar.html', consumivel=consumivel)

@consumiveis_bp.route('/excluir/<int:id>', methods=['DELETE'])
@login_required
def excluir(id):
    """Excluir consumível"""
    try:
        consumivel = Consumivel.query.get_or_404(id)
        codigo = consumivel.codigo
        filial = consumivel.filial
        
        db.session.delete(consumivel)
        db.session.commit()
        
        logger.info(f"Consumível {codigo} da filial {filial} excluído por {current_user.email}")
        
        return jsonify({
            'success': True,
            'message': 'Consumível excluído com sucesso!'
        })
    
    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro ao excluir consumível: {str(e)}")
        return jsonify({'error': f'Erro ao excluir consumível: {str(e)}'}), 500

@consumiveis_bp.route('/api/atualizar-estoque/<int:id>', methods=['POST'])
@login_required
def atualizar_estoque(id):
    """Atualizar apenas o estoque de um consumível"""
    try:
        consumivel = Consumivel.query.get_or_404(id)
        data = request.get_json()
        
        novo_saldo = int(data.get('saldo_estoque', consumivel.saldo_estoque))
        
        if novo_saldo < 0:
            return jsonify({'error': 'Saldo de estoque não pode ser negativo'}), 400
        
        consumivel.saldo_estoque = novo_saldo
        consumivel.atualizar_nivel_estoque()
        
        db.session.commit()
        
        logger.info(f"Estoque do consumível {consumivel.codigo} atualizado para {novo_saldo} por {current_user.email}")
        
        return jsonify({
            'success': True,
            'message': 'Estoque atualizado com sucesso!',
            'consumivel': consumivel.to_dict()
        })
    
    except ValueError as e:
        return jsonify({'error': 'Valor de estoque inválido'}), 400
    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro ao atualizar estoque: {str(e)}")
        return jsonify({'error': f'Erro ao atualizar estoque: {str(e)}'}), 500

@consumiveis_bp.route('/api/estatisticas')
@login_required
def estatisticas():
    """API para estatísticas de consumíveis"""
    try:
        total = Consumivel.query.count()
        critico = Consumivel.query.filter(Consumivel.nivel_estoque == 'Crítico').count()
        baixo = Consumivel.query.filter(Consumivel.nivel_estoque == 'Baixo').count()
        esgotado = Consumivel.query.filter(Consumivel.nivel_estoque == 'Esgotado').count()
        normal = Consumivel.query.filter(Consumivel.nivel_estoque == 'Normal').count()
        
        # Filiais com mais problemas
        filiais_problema = db.session.query(
            Consumivel.filial,
            db.func.count(Consumivel.id).label('total')
        ).filter(
            Consumivel.nivel_estoque.in_(['Crítico', 'Baixo', 'Esgotado'])
        ).group_by(Consumivel.filial).order_by(db.func.count(Consumivel.id).desc()).limit(5).all()
        
        return jsonify({
            'total': total,
            'critico': critico,
            'baixo': baixo,
            'esgotado': esgotado,
            'normal': normal,
            'filiais_problema': [{'filial': f[0], 'total': f[1]} for f in filiais_problema]
        })
    
    except Exception as e:
        logger.error(f"Erro ao buscar estatísticas: {str(e)}")
        return jsonify({'error': f'Erro ao buscar estatísticas: {str(e)}'}), 500

@consumiveis_bp.route('/exportar-excel')
@login_required
def exportar_excel():
    """Exportar consumíveis para Excel"""
    try:
        # Obter filtros da query string
        filial_filtro = request.args.get('filial', '')
        nivel_filtro = request.args.get('nivel', '')
        codigo_filtro = request.args.get('codigo', '')
        
        # Construir query com filtros
        query = Consumivel.query
        
        if filial_filtro:
            query = query.filter(Consumivel.filial.ilike(f'%{filial_filtro}%'))
        if nivel_filtro:
            query = query.filter(Consumivel.nivel_estoque == nivel_filtro)
        if codigo_filtro:
            query = query.filter(Consumivel.codigo.ilike(f'%{codigo_filtro}%'))
        
        # Buscar todos os consumíveis (sem paginação para exportação completa)
        consumiveis = query.order_by(Consumivel.filial, Consumivel.codigo).all()
        
        # Criar workbook Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Consumíveis"
        
        # Definir estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Cabeçalhos
        headers = [
            'ID', 'Filial', 'Código', 'Descrição', 'Unidade',
            'Saldo Estoque', 'Estoque Mínimo', 'Nível Estoque',
            'Data Cadastro', 'Data Atualização'
        ]
        
        # Aplicar estilo aos cabeçalhos
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Preencher dados
        for row, consumivel in enumerate(consumiveis, 2):
            ws.cell(row=row, column=1, value=consumivel.id)
            ws.cell(row=row, column=2, value=consumivel.filial)
            ws.cell(row=row, column=3, value=consumivel.codigo)
            ws.cell(row=row, column=4, value=consumivel.descricao)
            ws.cell(row=row, column=5, value=consumivel.unidade_medida)
            ws.cell(row=row, column=6, value=consumivel.saldo_estoque)
            ws.cell(row=row, column=7, value=consumivel.estoque_minimo)
            ws.cell(row=row, column=8, value=consumivel.nivel_estoque)
            ws.cell(row=row, column=9, value=consumivel.data_cadastro.strftime('%d/%m/%Y %H:%M') if consumivel.data_cadastro else '')
            ws.cell(row=row, column=10, value=consumivel.data_atualizacao.strftime('%d/%m/%Y %H:%M') if consumivel.data_atualizacao else '')
            
            # Aplicar bordas a todas as células
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = border
        
        # Ajustar largura das colunas
        column_widths = [8, 15, 15, 40, 12, 15, 15, 15, 20, 20]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Adicionar informações de filtros na planilha
        ws.cell(row=len(consumiveis) + 3, column=1, value="Filtros aplicados:")
        ws.cell(row=len(consumiveis) + 4, column=1, value=f"Filial: {filial_filtro or 'Todas'}")
        ws.cell(row=len(consumiveis) + 5, column=1, value=f"Nível: {nivel_filtro or 'Todos'}")
        ws.cell(row=len(consumiveis) + 6, column=1, value=f"Código: {codigo_filtro or 'Todos'}")
        ws.cell(row=len(consumiveis) + 7, column=1, value=f"Total de registros: {len(consumiveis)}")
        ws.cell(row=len(consumiveis) + 8, column=1, value=f"Exportado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        ws.cell(row=len(consumiveis) + 9, column=1, value=f"Exportado por: {current_user.email}")
        
        # Salvar em buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Criar resposta
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        # Nome do arquivo com timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'consumiveis_export_{timestamp}.xlsx'
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        logger.info(f"Exportação Excel realizada por {current_user.email} - {len(consumiveis)} registros")
        
        return response
        
    except Exception as e:
        logger.error(f"Erro ao exportar Excel: {str(e)}")
        flash('Erro ao exportar dados para Excel', 'error')
        return redirect(url_for('consumiveis.index'))

# ==================== ROTAS DE MOVIMENTAÇÃO DE ESTOQUE ====================

@consumiveis_bp.route('/movimentar/<int:id>', methods=['POST'])
@login_required
def movimentar_estoque(id):
    """Registrar movimentação de estoque (entrada ou saída)"""
    try:
        consumivel = Consumivel.query.get_or_404(id)
        data = request.get_json()
        
        tipo = data.get('tipo', '').upper()  # ENTRADA ou SAIDA
        quantidade = int(data.get('quantidade', 0))
        motivo = data.get('motivo', '').strip()
        observacao = data.get('observacao', '').strip()
        
        # Validações
        if tipo not in ['ENTRADA', 'SAIDA']:
            return jsonify({'error': 'Tipo de movimentação inválido. Use ENTRADA ou SAIDA'}), 400
        
        if quantidade <= 0:
            return jsonify({'error': 'Quantidade deve ser maior que zero'}), 400
        
        if not motivo:
            return jsonify({'error': 'Motivo é obrigatório'}), 400
        
        # Verificar se há saldo suficiente para saída
        if tipo == 'SAIDA' and quantidade > consumivel.saldo_estoque:
            return jsonify({
                'error': f'Saldo insuficiente. Disponível: {consumivel.saldo_estoque}'
            }), 400
        
        # Calcular novo saldo
        saldo_anterior = consumivel.saldo_estoque
        if tipo == 'ENTRADA':
            consumivel.saldo_estoque += quantidade
        else:
            consumivel.saldo_estoque -= quantidade
        
        saldo_atual = consumivel.saldo_estoque
        
        # Criar registro de movimentação
        movimentacao = MovimentacaoEstoque(
            consumivel_id=consumivel.id,
            tipo=tipo,
            quantidade=quantidade,
            saldo_anterior=saldo_anterior,
            saldo_atual=saldo_atual,
            motivo=motivo,
            observacao=observacao,
            usuario_id=current_user.id
        )
        
        # Atualizar nível de estoque
        consumivel.atualizar_nivel_estoque()
        
        db.session.add(movimentacao)
        db.session.commit()
        
        logger.info(f"Movimentação {tipo} de {quantidade} unidades no consumível {consumivel.codigo} por {current_user.email}")
        
        return jsonify({
            'success': True,
            'message': f'Movimentação registrada com sucesso!',
            'movimentacao': movimentacao.to_dict(),
            'consumivel': consumivel.to_dict()
        })
        
    except ValueError:
        return jsonify({'error': 'Quantidade inválida'}), 400
    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro ao registrar movimentação: {str(e)}")
        return jsonify({'error': f'Erro ao registrar movimentação: {str(e)}'}), 500


@consumiveis_bp.route('/historico/<int:id>')
@login_required
def historico_consumivel(id):
    """Buscar histórico de movimentações de um consumível"""
    try:
        consumivel = Consumivel.query.get_or_404(id)
        
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        
        movimentacoes = MovimentacaoEstoque.query.filter_by(
            consumivel_id=id
        ).order_by(
            MovimentacaoEstoque.data_movimentacao.desc()
        ).paginate(page=page, per_page=per_page, error_out=False)
        
        return jsonify({
            'consumivel': consumivel.to_dict(),
            'movimentacoes': [m.to_dict() for m in movimentacoes.items],
            'total': movimentacoes.total,
            'pages': movimentacoes.pages,
            'current_page': movimentacoes.page,
            'has_next': movimentacoes.has_next,
            'has_prev': movimentacoes.has_prev
        })
        
    except Exception as e:
        logger.error(f"Erro ao buscar histórico: {str(e)}")
        return jsonify({'error': f'Erro ao buscar histórico: {str(e)}'}), 500


@consumiveis_bp.route('/historico-geral')
@login_required
def historico_geral():
    """Buscar histórico geral de movimentações"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        
        # Filtros opcionais
        filial = request.args.get('filial', '')
        tipo = request.args.get('tipo', '')
        data_inicio = request.args.get('data_inicio', '')
        data_fim = request.args.get('data_fim', '')
        
        query = MovimentacaoEstoque.query.join(Consumivel)
        
        if filial:
            query = query.filter(Consumivel.filial.ilike(f'%{filial}%'))
        
        if tipo:
            query = query.filter(MovimentacaoEstoque.tipo == tipo.upper())
        
        if data_inicio:
            try:
                data_inicio_dt = datetime.strptime(data_inicio, '%Y-%m-%d')
                query = query.filter(MovimentacaoEstoque.data_movimentacao >= data_inicio_dt)
            except ValueError:
                pass
        
        if data_fim:
            try:
                data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d')
                data_fim_dt = data_fim_dt.replace(hour=23, minute=59, second=59)
                query = query.filter(MovimentacaoEstoque.data_movimentacao <= data_fim_dt)
            except ValueError:
                pass
        
        movimentacoes = query.order_by(
            MovimentacaoEstoque.data_movimentacao.desc()
        ).paginate(page=page, per_page=per_page, error_out=False)
        
        # Estatísticas
        total_entradas = MovimentacaoEstoque.query.filter_by(tipo='ENTRADA').count()
        total_saidas = MovimentacaoEstoque.query.filter_by(tipo='SAIDA').count()
        
        return jsonify({
            'movimentacoes': [m.to_dict() for m in movimentacoes.items],
            'total': movimentacoes.total,
            'pages': movimentacoes.pages,
            'current_page': movimentacoes.page,
            'has_next': movimentacoes.has_next,
            'has_prev': movimentacoes.has_prev,
            'estatisticas': {
                'total_entradas': total_entradas,
                'total_saidas': total_saidas
            }
        })
        
    except Exception as e:
        logger.error(f"Erro ao buscar histórico geral: {str(e)}")
        return jsonify({'error': f'Erro ao buscar histórico: {str(e)}'}), 500


@consumiveis_bp.route('/motivos-movimentacao')
@login_required
def motivos_movimentacao():
    """Retorna lista de motivos padrão para movimentações"""
    motivos = {
        'ENTRADA': [
            'Compra',
            'Transferência entrada',
            'Devolução',
            'Ajuste de inventário',
            'Doação recebida',
            'Outros'
        ],
        'SAIDA': [
            'Consumo interno',
            'Transferência saída',
            'Venda',
            'Descarte/Avaria',
            'Ajuste de inventário',
            'Empréstimo',
            'Outros'
        ]
    }
    return jsonify(motivos)
